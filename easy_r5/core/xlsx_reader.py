"""Standalone XLSX row reader — intended to be run as a subprocess.

Running inside the QGIS process is unsafe: Python's _elementtree.pyd C extension
conflicts with libxml2 loaded by QGIS's GDAL/Qt stack when called from a QgsTask
worker thread, causing a Windows fatal exception (access violation in
xmlDictReference).  Running in a fresh subprocess avoids all QGIS DLLs.

Usage:
    python xlsx_reader.py <json_args>

Args (JSON):
    {"path": "<absolute path>", "sheet": "<sheet name>" | null}

Output (stdout):
    {"rows": [[v, ...], ...], "sheet_names": [...], "sheet_used": "<name>"}
    or {"error": "<message>"}

Cell values are serialised to JSON-compatible types:
    int / float / str / bool / None → unchanged
    anything else (e.g. datetime)  → str(v)
"""

from __future__ import annotations

import json
import sys


def _serialize(v: object) -> object:
    if v is None or isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


def read_xlsx(path: str, sheet: str | None) -> dict:
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_names: list[str] = wb.sheetnames

    if sheet:
        if sheet not in sheet_names:
            wb.close()
            return {
                "error": "Sheet '{}' not found in '{}'. Available sheets: {}.".format(
                    sheet, path, ", ".join(sheet_names)
                )
            }
        ws = wb[sheet]
        sheet_used = sheet
    else:
        ws = wb.worksheets[0]
        sheet_used = sheet_names[0]

    rows = [[_serialize(cell.value) for cell in row] for row in ws.iter_rows()]
    wb.close()

    return {"rows": rows, "sheet_names": sheet_names, "sheet_used": sheet_used}


if __name__ == "__main__":
    args = json.loads(sys.argv[1])
    try:
        result = read_xlsx(args["path"], args.get("sheet"))
    except Exception as exc:  # noqa: BLE001
        result = {"error": str(exc)}
    print(json.dumps(result, ensure_ascii=False))
